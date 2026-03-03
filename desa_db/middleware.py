import os
import json
import re
import duckdb
import html as h
import polars as pl
import openpyxl
import csv
import io
import time
import traceback
import gc
from difflib import SequenceMatcher
from fastapi import Response
from fastapi.responses import JSONResponse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# CONFIGURATION
# ==========================================
# /root
#   /root
#   /.config/headers.json
#   /.config/intervensi_kegiatan_mapping.json
#   /.config/rekomendasi.json
#   /.config/table_structure.csv
#   /desa_db/server.py
#   /desa_db/middleware.py
#   /front_end/router.py
#   /front_end/templates/admin.html
#   /front_end/templates/user.html
#   /front_end/templates/login.html


# Resolves to: /.../desa_db/
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Resolves to: /.../.config/
CONFIG_DIR = os.path.abspath(os.path.join(BASE_DIR, "../.config"))

# Resolves to: /.../desa_db/dbs/
DB_FOLDER = os.path.join(BASE_DIR, "dbs")

# Resolves to: /.../desa_db/temp/
TEMP_FOLDER = os.path.join(BASE_DIR, "temp")
os.makedirs(TEMP_FOLDER, exist_ok=True)

# Config Files
JSON_PATH = os.path.join(CONFIG_DIR, "rekomendasi.json")

# intervensi_kegiatan_mapping files
INTERVENTION_FILE = os.path.join(CONFIG_DIR, "intervensi_kegiatan_mapping.json")

# Header File: inside .config/headers.json (sibling of desa_db)
HEADER_FILE = os.path.join(CONFIG_DIR, "headers.json")

# Data Structure Constants
# ID_COL acts as the Primary Key for deduplication and version control.
ID_COL = "Kode Wilayah Administrasi Desa" 

# Export Folder for Ready to download
EXPORT_FOLDER = os.path.abspath(os.path.join(BASE_DIR, "../exports"))
os.makedirs(EXPORT_FOLDER, exist_ok=True)

# Point to ../.config/rekomendasi.json relative to this file
os.makedirs(DB_FOLDER, exist_ok=True)

# ==========================================
# LOGIC LOADER (Existing)
# ==========================================
def load_logic():
    """
    Loads rekomendasi.json and converts string keys ("1", "2", ...) to integers.
    Returns empty dict if file is missing (safe fallback).
    """
    if not os.path.exists(JSON_PATH):
        # Fallback: Return empty dict if file is missing (prevents crashes)
        return {}

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        raw_data = json.load(f)

    # CRITICAL STEP: Convert JSON String keys ("1") to Integers (1)
    # If we don't do this, Polars won't match the Integer data in DB.
    clean_logic = {}
    for col_name, mapping in raw_data.items():
        # Convert keys to int, keep values as is
        clean_logic[col_name] = {int(k): v for k, v in mapping.items()}

    return clean_logic

# Load once on startup
RECOMMENDATION_LOGIC = load_logic()

# ==========================================
# OPTIMIZED BATCH EXECUTION
# ==========================================
def apply_rekomendasis(df: pl.DataFrame) -> pl.DataFrame:
    """
    Applies all recommendation rules from RECOMMENDATION_LOGIC in a single batch pass.
    Uses replace_strict for Polars 1.0+ compatibility and optimal performance.
    Only processes columns that exist in the DataFrame.
    """
    exprs = []

    # Build the list of instructions (Lazy Evaluation)
    for col_name, logic_map in RECOMMENDATION_LOGIC.items():
        if col_name in df.columns:
            expr = (
                pl.col(col_name)
                # OPTIMIZATION: Data is already Int from DB.
                # This explicitly handles the 'default' and 'return_dtype' logic
                .replace_strict(logic_map, default=None, return_dtype=pl.String)
                .alias(col_name)
            )
            exprs.append(expr)

    # Execute all instructions at once
    if exprs:
        return df.with_columns(exprs)

    return df

def make_json_response(df: pl.DataFrame) -> Response:
    """
    OPTIMIZATION: 
    Converts Polars DataFrame directly to JSON string using Rust engine
    and returns a FastAPI Response. Bypasses Python's slow list-of-dicts conversion
    for better performance on large datasets.
    """
    # Serialize directly to JSON string using Rust engine
    json_str = df.write_json()

    # Return direct Response (skips FastAPI's jsonable_encoder overhead)
    return Response(content=json_str, media_type="application/json")

# ==========================================
# DATABASE HELPERS
# ==========================================

# PREVIEW & MAPPING HELPERS, sends back to frontend for user to check headers
def helpers_read_excel_preview(file_path: str, limit_rows: int = 25):
    """
    Reads first N rows of 'Skor' sheet for frontend preview.
    - .xlsx: openpyxl streaming (read_only=True) for low memory
    - .xlsb/.xlsx fallback: Polars calamine
    - Caps at 400 columns
    - Converts None → ""
    - Returns pure list-of-lists
    Returns:
        list[list[str]]: rows as list of lists (max limit_rows rows, 400 cols)
    """
    rows = []

    try:
        # OPTIMIZATION: Streaming for .xlsx
        if file_path.lower().endswith(".xlsx"):
            # read_only=True ensures we don't load the whole file
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            try:
                # Handle Sheet Selection
                if "Skor" in wb.sheetnames:
                    ws = wb["Skor"]
                else:
                    ws = wb.active

                # Iterate only the rows we need
                # fetch limit_rows + 5 just in case some are empty, but typically exact limit is fine
                for row in ws.iter_rows(min_row=1, max_row=limit_rows, values_only=True):
                    # Convert None to "" and map to list
                    clean_row = [str(cell) if cell is not None else "" for cell in row]
                    
                    # Check if row has data
                    rows.append(clean_row[:400])

            finally:
                wb.close()

            return rows

        # Read without headers to get raw grid
        else:
            df = pl.read_excel(
                file_path,
                sheet_name="Skor",
                engine="calamine",
                has_header=False # Direct argument, not inside dict
            )
            
            # Slice row limit manually
            df = df.head(limit_rows)

            # Limit width (similar to existing logic)
            max_col = min(df.width, 400)
            df = df[:, :max_col]
            
            # Replace nulls with empty string for JSON safety
            data = df.fill_null("").to_dicts()
            
            # Convert to list of lists (easier for generic grid in JS)
            # Polars to_dicts returns list of objects with "column_0", "column_1", etc.
            # Only pure values.
            rows = [list(row.values()) for row in data]
            return rows

    except Exception as e:
        raise Exception(f"Preview Error: {str(e)}")

def helpers_normalize_text(text: str):
    """
    Normalizes text for fuzzy header matching:
    - Lowercase
    - Remove quotes
    - Collapse multiple whitespace
    - Strip
    """
    if not text: return ""
    text = str(text).lower()
    text = re.sub(r'["\']', '', text) # Remove quotes
    text = re.sub(r'\s+', ' ', text)  # Collapse spaces
    return text.strip()

def helpers_generate_header_mapping(file_path: str, header_row_idx: int):
    """
    Generates column mapping from uploaded Excel against headers.json.
    
    Logic:
    - Loads standards and aliases from headers.json
    - .xlsx: streams only header row with openpyxl (read_only=True)
    - .xlsb: full read with calamine + row extraction
    - Caps at 400 columns
    - Detects first non-empty column → start_col (dynamic alignment)
    - First 6 columns: forced index mapping + auto-confirm
    - Remaining columns: two-pass matching (skips empty headers)
      - Pass 1: alias_exact (exact match on standard or aliases)
      - Pass 2: fuzzy SequenceMatcher on leftovers
        - > 85% → auto-confirm
        - < 40% → "(No Match)"
    - Tracks assigned_standards to prevent duplicates
    - Computes missing_headers (unmapped standards after first 6)
    
    Returns:
        tuple[list[dict], list[str]]: (
            mapping: list of dicts with keys: standard, file_header,
                     col_index, is_confirmed, match_type, score,
            missing_headers: list of unmapped standard names
        )
    """

    # Load Standard Headers
    if not os.path.exists(HEADER_FILE):
        raise Exception("Configuration file not found: headers.json")

    with open(HEADER_FILE, "r", encoding="utf-8") as f:
        headers_data = json.load(f)
        standard_headers = [item["standard"] for item in headers_data]

    file_headers_raw = []

    # OPTIMIZATION: Streaming for .xlsx
    if file_path.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if "Skor" in wb.sheetnames:
            ws = wb["Skor"]
        else:
            ws = wb.active
        
        # OpenPyXL is 1-based, header_row_idx is 0-based.
        target_excel_row = header_row_idx + 1
        
        # Iterate until we hit the row (efficient skipping)
        for i, row in enumerate(ws.iter_rows(min_row=target_excel_row, max_row=target_excel_row, values_only=True)):
            clean_row = [str(cell) if cell is not None else "" for cell in row]
            file_headers_raw = clean_row[:400]
            break
        
        wb.close()
        
        if not file_headers_raw:
             raise Exception(f"Header row index {header_row_idx} is out of bounds or empty.")
             
    else:
        # Fallback for .xlsb
        df = pl.read_excel(
            file_path,
            sheet_name="Skor",
            engine="calamine",
            has_header=False
        )
        max_col = min(df.width, 400)
        df = df[:, :max_col]

        if header_row_idx >= df.height:
            raise Exception(f"Header row index {header_row_idx} is out of bounds")

        # Get the raw list of values from the target row
        # Polars reads as col_0, col_1.
        file_headers_raw = list(df.row(header_row_idx))
    
    mapping = []
    special_first_6 = standard_headers[:6]
    remaining_standards = standard_headers[6:]
    
    # DYNAMIC ALIGNMENT: Find the first non-empty column in the Excel file.
    # This prevents assigning "Provinsi" to an empty Column A.
    start_col = 0
    for i, val in enumerate(file_headers_raw):
        if val.strip() != "":
            start_col = i
            break

    # Set to track already assigned header
    assigned_standards = set()

    # Map Special First 6 using the actual start column
    for i in range(min(6, len(file_headers_raw) - start_col)):
        if start_col + i >= len(file_headers_raw):
            break
        file_val = str(file_headers_raw[start_col + i]).strip()
        std_val = special_first_6[i]
        
        # Auto-confirm the first 6 based on User's rule "the order always right"
        mapping.append({
            "standard": std_val,
            "file_header": file_val,
            "col_index": start_col + i,
            "is_confirmed": True, 
            "match_type": "index_forced"
        })
        assigned_standards.add(std_val)

    # --- TWO PASS MATCHING ALGORITHM ---
    mapping_dict = {}

    # Resume fuzzy mapping after the first 6
    fuzzy_start_idx = start_col + 6

    # PASS 1: EXACT & ALIAS MATCHES ONLY
    for i in range(fuzzy_start_idx, len(file_headers_raw)):
        file_val = str(file_headers_raw[i]).strip()

        # Even if empty, we list it so user can see it (but usually ignore)
        if not file_val: continue # Skip empty columns
        
        # SKIP EMPTY COLUMNS
        # If the header in the file is empty/None/NaN, it's not a valid data column.
        # Don't even try to map it.
        if not file_val or file_val.lower() == "none" or file_val == "":
            continue

        norm_file = helpers_normalize_text(file_val)
        best_match = None

        # Match based on headers.json and it's aliases
        for item in headers_data[6:]:
            original_std = item["standard"]
            
            # Skip if this standard header has already been mapped to an earlier column
            if original_std in assigned_standards:
                continue

            valid_names = [helpers_normalize_text(a) for a in item.get("aliases", []) if a.strip()]
            valid_names.append(helpers_normalize_text(original_std))

            if norm_file in valid_names:
                best_match = original_std
                break

        if best_match:
            assigned_standards.add(best_match)
            mapping_dict[i] = {
                "standard": best_match,
                "file_header": file_val,
                "col_index": i,
                "is_confirmed": True,
                "match_type": "alias_exact",
                "score": 1.0
            }
        else:
            # Leave for Pass 2
            mapping_dict[i] = {
                "standard": None,
                "file_header": file_val,
                "col_index": i,
                "is_confirmed": False,
                "match_type": "fuzzy",
                "score": 0.0
            }

    # PASS 2: FUZZY MATCH THE LEFTOVERS
    norm_standards = [(helpers_normalize_text(h), h) for h in remaining_standards]
    
    for i in range(fuzzy_start_idx, len(file_headers_raw)):
        if i not in mapping_dict:
            continue
            
        if mapping_dict[i]["standard"] is None:
            file_val = mapping_dict[i]["file_header"]
            norm_file = helpers_normalize_text(file_val)
            
            best_match = None
            highest_ratio = 0.0
            
            for norm_std, orig_std in norm_standards:
                # Skip if already mapped
                if orig_std in assigned_standards:
                    continue

                # Fuzzy match (Levenshtein) using python's difflib
                ratio = SequenceMatcher(None, norm_std, norm_file).ratio()
                if ratio > highest_ratio:
                    highest_ratio = ratio
                    best_match = orig_std

            # If the best match is less than 40%, it's likely garbage.
            if highest_ratio < 0.4:
                best_match = None
            # Lock the standard header so it cannot be assigned again
            if best_match:
                assigned_standards.add(best_match)
            mapping_dict[i]["standard"] = best_match if best_match else "(No Match)"
            mapping_dict[i]["score"] = round(highest_ratio, 2)
            mapping_dict[i]["is_confirmed"] = highest_ratio > 0.85
            
    # Rebuild mapping array in left-to-right order
    for i in range(fuzzy_start_idx, len(file_headers_raw)):
        if i in mapping_dict:
            mapping.append(mapping_dict[i])

    # Compute missing headers (excluding the first 6 fixed columns)
    found_standards = {m["standard"] for m in mapping if m["standard"] != "(No Match)"}
    missing_headers = [h for h in remaining_standards if h not in found_standards]

    return mapping, missing_headers

# PROCESSOR to rebould db based on headers.json
def helpers_sync_db_schema(con: duckdb.DuckDBPyConnection, year: str):
    """
    Synchronizes master_data table schema with headers.json.
    Adds any missing columns (VARCHAR for text columns, TINYINT for scores).
    """
    if not os.path.exists(HEADER_FILE): return

    with open(HEADER_FILE, "r", encoding="utf-8") as f:
        headers_data = json.load(f)
        target_headers = [item["standard"] for item in headers_data]

    # Get existing DB columns
    existing_cols = set([r[0] for r in con.execute("DESCRIBE master_data").fetchall()])
    
    TEXT_COLUMNS = {
        "Provinsi", "Kabupaten/ Kota", "Kecamatan", 
        "Kode Wilayah Administrasi Desa", "Desa", "Status ID"
    }

    for h in target_headers:
        # If header contains quotes, strip them for comparison
        clean_h = h.replace('"', '')
        if clean_h not in existing_cols:
            print(f"Schema Evolution: Adding column '{clean_h}' to {year}...")
            # Determine Type
            dtype = "VARCHAR" if clean_h in TEXT_COLUMNS else "TINYINT"
            try:
                con.execute(f'ALTER TABLE master_data ADD COLUMN "{clean_h}" {dtype}')
            except Exception as e:
                print(f"Failed to add column {clean_h}: {e}")

# INTERNAL FILE PROCESSOR, staging uploads basically
def helpers_internal_process_temp_file(
    year: str,
    file_path: str,
    filename: str,
    temp_id: str,
    header_row_idx: int,
    data_start_idx: int,
    confirmed_mapping: list
    ):
    """
    Full ETL processor for staging upload.
    
    Alignment fixes:
    - openpyxl scan: counts leading empty rows + first non-empty column offset
    - calamine read (has_header=False) then applies row/col offsets
    - Duplicate column resolution: prefers column with highest numeric values
    
    Processing steps:
    - Row slice after offset
    - Confirmed mapping only (renames to standard names)
    - Drop fully null/empty columns
    - Strict Status ID validation
    - Filter 10-digit IDs
    - Cast: text columns → String, scores → Int8 (strict=False)
    - Write {temp_id}.parquet
    - CDC diff vs current master_data (valid_to IS NULL)
    Returns:
        dict: {
            "status": "staged",
            "temp_id": str,
            "filename": str,
            "diff": {
                "added": int,
                "removed": int,
                "changed": int,
                "total_incoming": int
            }
        }
    """

    try:

        # --- ETL: Extraction & Cleaning (Polars) ---
        print(f"Processing {filename} for year {year}...")

        # ALIGNMENT FIX: Count ALL empty rows up to the user's selected start index.
        # This perfectly syncs the frontend openpyxl index with the backend Calamine index
        # regardless of how many empty rows exist between headers and data.
        calamine_row_offset = 0
        calamine_col_offset = 0
        
        if file_path.lower().endswith(".xlsx"):
            try:
                wb_check = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws_check = wb_check["Skor"] if "Skor" in wb_check.sheetnames else wb_check.active
                
                empty_rows_count = 0
                min_col = 9999
                
                # Check up to data_start_idx to count every empty row Calamine might drop
                for r_idx, row in enumerate(ws_check.iter_rows(min_row=1, max_row=data_start_idx + 1, values_only=True)):
                    row_has_data = False
                    for c_idx, cell in enumerate(row):
                        if cell is not None and str(cell).strip() != "":
                            row_has_data = True
                            if c_idx < min_col:
                                min_col = c_idx
                    
                    if not row_has_data:
                        empty_rows_count += 1
                
                calamine_row_offset = empty_rows_count
                calamine_col_offset = min_col if min_col != 9999 else 0
                wb_check.close()
            except Exception as e:
                print(f"Warning calculating offset: {e}")

        # Read .xlsb/.xlsx using 'calamine'
        df = pl.read_excel(
            file_path,
            sheet_name="Skor",
            engine="calamine",
            has_header=False # Direct argument, not inside dict
        )
        # DEBUG TAG
        print(
        f"1. Initial read: {df.height} rows (Empty rows stripped: {calamine_row_offset}, Col Offset: {calamine_col_offset})",
        flush=True
        )

        # Structure Cleaning:
        # Slice: Cap width to prevent reading infinite empty Excel columns
        max_col = min(df.width, 400) 
        df = df[:, :max_col] # DO NOT drop columns manually anymore; Calamine offset handles it safely

        # Sync the absolute UI index with Calamine's stripped index
        adjusted_data_start = max(0, data_start_idx - calamine_row_offset)

        # Slice Rows (Data Start)
        # Assume data_start_idx is the 0-based index from the file where actual data begins.
        # Example: Header at 3, Data at 4.
        # df[4] is the first data row.
        if adjusted_data_start < df.height:
            df = df.slice(adjusted_data_start, None)
            # DEBUG TAG
            print(
            f"2. After row slice (UI start={data_start_idx}, Adjusted={adjusted_data_start}): {df.height} rows",
            flush=True
            )

        else:
             raise Exception(f"Data start index {data_start_idx} is out of bounds")

        # Apply Column Mapping
        target_to_cols = {}
        for m in confirmed_mapping:
            if m.get('is_confirmed'):
                t = m['standard']
                if t not in target_to_cols:
                    target_to_cols[t] = []

                # Adjust absolute UI index to relative Calamine index
                cal_idx = m['col_index'] - calamine_col_offset
                if cal_idx >= 0:
                    target_to_cols[t].append(cal_idx)

        selection_exprs = []
        for target_name, indices in target_to_cols.items():
            if not indices:
                continue
                
            chosen_idx = indices[0]

            # Resolve duplicates by checking if the dataframe column contains numbers
            if len(indices) > 1:
                for idx in indices:
                    if idx < df.width:
                        col_name = df.columns[idx]
                        # Cast to Int8 (strict=False turns text/invalid data to null). 
                        # If it has valid numbers, it is the correct score column.
                        numeric_count = df[col_name].cast(pl.Int8, strict=False).is_not_null().sum()
                        if numeric_count > 0:
                            chosen_idx = idx
                            break 

            if 0 <= chosen_idx < df.width:
                original_name = df.columns[chosen_idx]
                selection_exprs.append(pl.col(original_name).alias(target_name))

        if not selection_exprs:
            raise Exception("No columns were confirmed for mapping.")

        df = df.select(selection_exprs)
        # DEBUG TAG
        print(f"3. After column mapping: {df.height} rows", flush=True)
        if ID_COL in df.columns:
            print(f" -> Sample data in ID_COL: {df[ID_COL].head(3).to_list()}", flush=True)

        # Filter: Remove columns that are entirely Empty or Null
        df = df.select([
            col for col in df 
            if not (col.is_null().all() or (col.dtype == pl.String and (col == "").all()))
        ])
        # DEBUG TAG
        print(f"4. After dropping empty cols: {df.height} rows", flush=True)

        # VALIDATION: Status ID (Strict)
        if "Status ID" in df.columns:
            valid_statuses = ["BERKEMBANG", "MAJU", "MANDIRI", "SANGAT TERTINGGAL", "TERTINGGAL"]
            
            # Check for invalid values
            # is_in returns boolean series. ~ negates it.
            invalid_rows = df.filter(
                ~pl.col("Status ID").str.to_uppercase().is_in(valid_statuses) 
                & pl.col("Status ID").is_not_null() 
                & (pl.col("Status ID") != "")
            )
            
            if invalid_rows.height > 0:
                bad_vals = invalid_rows["Status ID"].unique().to_list()
                raise Exception(f"Invalid 'Status ID' found: {bad_vals}. Allowed: {valid_statuses}")

        # Filter: Keep only rows where at least one column looks like an ID
        # Pattern: 10 digits (Standard Indonesian Region Code)
        if ID_COL in df.columns:
            pattern = r"[0-9]{10}"
            match_count = df.filter(pl.col(ID_COL).cast(pl.String).str.contains(pattern)).height
            # DEBUG TAG
            print(f"5. Rows matching exactly 10 digits: {match_count}", flush=True)
            
            df = df.filter(
                pl.col(ID_COL).cast(pl.String).str.contains(pattern)
            )
            # DEBUG TAG
            print(f"6. After ID filter: {df.height} rows", flush=True)

        # OPTIMIZATION Type Casting (Optimized)
        # Identify Text Columns (same as middleware)
        TEXT_COLUMNS = {
            "Provinsi", "Kabupaten/ Kota", "Kecamatan", 
            "Kode Wilayah Administrasi Desa", "Desa", "Status ID"
        }

        # Cast, if non TEXT_COLUMNS return int8 (TINYINT)
        exprs = []
        for col in df.columns:
            if col in TEXT_COLUMNS:
                exprs.append(pl.col(col).cast(pl.String))
            else:
                exprs.append(pl.col(col).cast(pl.Int8, strict=False))
        
        df = df.with_columns(exprs)

        # Save Staged Data (Parquet will now store these as Integers)
        final_parquet_path = os.path.join(TEMP_FOLDER, f"{temp_id}.parquet")
        df.write_parquet(final_parquet_path)

        # Uncomment for the debug to activate. what this debug does is to check what's inside
        # the excel files
        # print(f"\n--- DEBUG LOG FOR {filename} ---", flush=True)
        # print(f"Total rows remaining after Polars cleaning: {df.height}", flush=True)
        # 
        # if ID_COL in df.columns:
        #     sample_ids = df[ID_COL].head(20).to_list()
        #     print(f"Sample IDs (First 20): {sample_ids}", flush=True)
        # else:
        #     print(f"CRITICAL WARNING: Column '{ID_COL}' is MISSING!", flush=True)
        # print("--------------------------------\n", flush=True)

        # Compare with DB (CDC Logic - Identical to previous)
        con, db_path = helpers_get_db_connection(year)
        helpers_init_db(con, df.columns) # Ensure tables exist
        helpers_sync_db_schema(con, year) # Check headers.json vs DB

        # Identify Incoming Data
        con.execute(f"CREATE OR REPLACE TEMP TABLE incoming AS SELECT * FROM read_parquet('{final_parquet_path}')")
        # Current Active State
        con.execute("CREATE OR REPLACE TEMP TABLE current_state AS SELECT * FROM master_data WHERE valid_to IS NULL")

        # Rows to ADD (ID in Incoming, Not in Current)
        added_count = con.execute(f"""
            SELECT COUNT(*) FROM incoming i 
            WHERE i."{ID_COL}" NOT IN (SELECT "{ID_COL}" FROM current_state)
        """).fetchone()[0]

        # Rows to REMOVE (ID in Current, Not in Incoming) - FULL SYNC OPTION
        removed_count = con.execute(f"""
            SELECT COUNT(*) FROM current_state c
            WHERE c."{ID_COL}" NOT IN (SELECT "{ID_COL}" FROM incoming)
        """).fetchone()[0]

        # Rows to UPDATE (ID Match, Content Differs)
        other_cols = [h for h in df.columns if h != ID_COL]
        # Handle single column case or empty other_cols
        if not other_cols:
            changed_count = 0
        else:
            check_conditions = " OR ".join(['(c."{col}" IS DISTINCT FROM i."{col}")'.format(col=h.replace('"', '""')) for h in other_cols])
            changed_count = con.execute(f"""
                SELECT COUNT(*) FROM current_state c
                JOIN incoming i ON c."{ID_COL}" = i."{ID_COL}"
                WHERE {check_conditions}
            """).fetchone()[0]

        con.close()

        return {
            "status": "staged",
            "temp_id": temp_id,
            "filename": filename,
            "diff": {
                "added": added_count,
                "removed": removed_count,
                "changed": changed_count,
                "total_incoming": df.height
            }
        }
    except Exception as e:
        raise e

def helpers_get_db_connection(year: str):
    """
    Returns DuckDB connection and path for the given year.
    Validates year format (must be 20XX).
    """
    clean_year = os.path.basename(year)

    if not re.match(r'^20\d{2}$', clean_year):
        raise ValueError(f"Year must start with 20XX, got: '{clean_year}'")

    db_path = os.path.join(DB_FOLDER, f"data_{clean_year}.duckdb")
    con = duckdb.connect(db_path)

    return con, db_path

def helpers_get_cache_path(year: str):
    """
    Returns the absolute path to the unique_cache JSON file for a specific year.
    """
    get_cache_path_clean_year = os.path.basename(year)
    return os.path.join(CONFIG_DIR, f"unique_cache_{get_cache_path_clean_year}.json")

def helpers_init_db(con: duckdb.DuckDBPyConnection, headers: list[str]):
    """
    SCD Type 2 Schema:

    Initializes SCD Type 2 schema for master_data and commits tables.
    - Creates tables if they don't exist
    - Uses TINYINT for score columns, VARCHAR for text/metadata
    - Adds index on ID_COL
    """
    # Construct column definitions for SQL (e.g., "Col1" VARCHAR, "Col2" VARCHAR...)
    # Note: Wrap headers in double quotes to handle spaces or special chars in column names.

    # define columns that MUST remain text/string
    TEXT_COLUMNS = {
        "valid_from", "valid_to", "commit_id", "source_file",
        "Provinsi", "Kabupaten/ Kota", "Kecamatan", 
        "Kode Wilayah Administrasi Desa", "Desa", "Status ID"
    }

    cols_def = []
    for h in headers:
        if h in TEXT_COLUMNS:
            cols_def.append(f'"{h}" VARCHAR')
        else:
            # OPTIMIZATION: Use TINYINT for 1-5 scores
            cols_def.append(f'"{h}" TINYINT')

    cols_optimized = ", ".join(cols_def)

    # Create the MAIN table (Latest Snapshot)
    # Includes metadata: 'last_updated' and 'source_file'
    con.execute(f"""
        CREATE TABLE IF NOT EXISTS master_data (
            valid_from TIMESTAMP,
            valid_to TIMESTAMP,
            commit_id VARCHAR,
            source_file VARCHAR,
            {cols_optimized}
        )
    """)

    # Index for speed
    con.execute(f"CREATE INDEX IF NOT EXISTS idx_id ON master_data (\"{ID_COL}\")")

    # Create the HISTORY table (Audit Log)
    # Identical structure but includes 'archived_at' to track when the record was replaced.
    con.execute("""
        CREATE TABLE IF NOT EXISTS commits (
            commit_id VARCHAR PRIMARY KEY,
            timestamp TIMESTAMP,
            filename VARCHAR,
            summary VARCHAR
        )
    """)

def helpers_build_dynamic_query(con, base_query, request_params, base_values=None):
    """
    Builds parameterized WHERE clause from query params.
    Supports:
    - ids (semicolon/newline separated) → IN (?)
    - column ILIKE ? for any valid column
    - filterModel (AG Grid JSON) for server-side column sorting & filtering
    - Prepends base_values (e.g. for time-travel ?version)
    Returns:
        tuple[str, list]: (final_query_with_where, values_list_for_execution)
    """
    # Get valid column names from the 'master_data' table to prevent SQL injection
    valid_cols = [r[0] for r in con.execute("DESCRIBE master_data").fetchall()]

    filters = []
    values = base_values[:] if base_values else []

    # Iterate over all query parameters
    # A. SPECIAL FILTER: ID List (Semicolon Separated)
    if "ids" in request_params and request_params["ids"]:
        raw_ids = request_params["ids"]

        # STRICT RULE: Split by SEMICOLON (;) or Newline. NO COMMAS.
        filtered_ids = re.split(r'[;\n\r]+', raw_ids)

        # Strip each token and remove empty strings in one pass
        id_list = [x.strip() for x in filtered_ids if x.strip()]

        if id_list:
            # Create placeholders (?, ?, ?)
            placeholders = ", ".join(["?"] * len(id_list))
            filters.append(f'"{ID_COL}" IN ({placeholders})')
            values.extend(id_list)

    # B. AG GRID DYNAMIC COLUMN FILTERS
    if "filterModel" in request_params and request_params["filterModel"]:
        try:
            filter_model = json.loads(request_params["filterModel"])
            for col, config in filter_model.items():
                if col in valid_cols:
                    f_type = config.get("filterType", "text")
                    match_type = config.get("type", "contains")
                    val = config.get("filter")
                    
                    if val is not None:
                        if match_type == "contains":
                            filters.append(f'"{col}" ILIKE ?')
                            values.append(f"%{val}%")
                        elif match_type == "notContains":
                            filters.append(f'"{col}" NOT ILIKE ?')
                            values.append(f"%{val}%")
                        elif match_type == "equals":
                            if f_type == "number":
                                filters.append(f'"{col}" = ?')
                                values.append(val)
                            else:
                                filters.append(f'"{col}" ILIKE ?')
                                values.append(val)
                        elif match_type == "notEqual":
                            if f_type == "number":
                                filters.append(f'"{col}" != ?')
                                values.append(val)
                            else:
                                filters.append(f'"{col}" NOT ILIKE ?')
                                values.append(val)
                        elif match_type == "startsWith":
                            filters.append(f'"{col}" ILIKE ?')
                            values.append(f"{val}%")
                        elif match_type == "endsWith":
                            filters.append(f'"{col}" ILIKE ?')
                            values.append(f"%{val}")
                        elif match_type == "greaterThan":
                            filters.append(f'CAST("{col}" AS DOUBLE) > ?')
                            values.append(float(val))
                        elif match_type == "lessThan":
                            filters.append(f'CAST("{col}" AS DOUBLE) < ?')
                            values.append(float(val))
                        elif match_type == "greaterThanOrEqual":
                            filters.append(f'CAST("{col}" AS DOUBLE) >= ?')
                            values.append(float(val))
                        elif match_type == "lessThanOrEqual":
                            filters.append(f'CAST("{col}" AS DOUBLE) <= ?')
                            values.append(float(val))
        except Exception as e:
            print(f"Filter parsing error: {e}")

    # C. STANDARD FILTERS
    # Ignore filterModel here so it doesn't trigger the basic matching logic below
    ignored_keys = ["ids", "version", "limit", "translate", "temp_id", "filename", "sort_by", "sort_dir", "filterModel"]
    for key, val in request_params.items():
        # Map URL friendly names if necessary, or use direct keys
        # Assume frontend sends specific column names like "Provinsi"
        if key in valid_cols and key not in ignored_keys and val:
            # Handle the explicit "NONE" tag (user unchecked everything)
            if val == "__NONE__":
                filters.append("1=0")
                continue
            
            # Handle multi-select IN clause
            if ";" in val:
                tokens = [x.strip() for x in val.split(";") if x.strip()]
                if tokens:
                    placeholders = ", ".join(["?"] * len(tokens))
                    filters.append(f'"{key}" IN ({placeholders})')
                    values.extend(tokens)
            else:
                # Fallback for standard single-text search
                filters.append(f'"{key}" ILIKE ?')
                values.append(f"%{val}%")

    # Smart Append
    if filters:
        filter_str = " AND ".join(filters)
        # Check if base_query already has a WHERE clause (case-insensitive)
        if "WHERE" in base_query.upper():
            base_query += f" AND {filter_str}"
        else:
            base_query += f" WHERE {filter_str}"

    return base_query, values

def helpers_generate_excel_workbook(con, df_grid: pl.DataFrame, do_translate: bool, params_dict: dict = None) -> Workbook:
    """
    Builds a complete three-sheet Excel Workbook from the provided DataFrame.
    
    Sheet 1: "Grid Data"
    - Full rows from df_grid
    - If do_translate=True: applies apply_rekomendasis() (scores → text)
    - Simple header + data + basic bold/fill styling
    
    Sheet 2: "Dashboard Rekomendasi"
    - Loads table_structure.csv
    - Determines metric column order from master_data schema
    - Loads intervention templates
    - Calls helpers_calculate_dashboard_stats() for averages, counts, and narrative
    - Builds fully formatted table with:
        • Proper merged headers (SKOR group + PELAKSANA group)
        • Column widths optimized for readability
        • Rowspans for NO / DIMENSI / SUB DIMENSI / INDIKATOR
        • Borders, alignment, and styling
    
    Sheet 3: "Dashboard IKU"
    - Loads table_structure_IKU.csv and iku_mapping.json
    - Determines grouping level from params_dict (Provinsi → Kabupaten → Kecamatan → Desa)
    - Maps CSV headers: parents (merged colspans) + subs (statuses, rata-rata, total, capaian)
    - Computes per-parent IKU scores: averages over mapped children columns
    - Aggregates by group: JLH DESA, averages, status counts (≥4 threshold)
    - Adds TOTAL row with sums/averages
    - Appends grouped data rows
    - Applies borders, number formats (#,##0 for int; #,##0.00 for float)
    
    Returns:
        openpyxl.workbook.workbook.Workbook: fully built workbook (not saved)
    """
    # BUILD EXCEL (OpenPyXL)
    wb = Workbook()
    
    # Define Common Styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    align_top = Alignment(vertical='top', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True)

    # --- SHEET 1: GRID DATA ---
    ws1 = wb.active
    ws1.title = "Grid Data"
    df_sheet1 = apply_rekomendasis(df_grid) if do_translate else df_grid
    grid_dicts = df_sheet1.to_dicts()
    
    # Write Headers
    if grid_dicts:
        headers = list(grid_dicts[0].keys())
        ws1.append(headers)
        for row in grid_dicts:
            # Handle None/NaN for Excel
            clean_row = [(v if v is not None else "") for v in row.values()]
            ws1.append(clean_row)
        
        # Basic Formatting for Sheet 1
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
    else:
        ws1.append(["No Data Found for filters"])

    del grid_dicts  # free after writing to workbook
    gc.collect()

    # --- SHEET 2: DASHBOARD REKOMENDASI ---
    ws2 = wb.create_sheet("Dashboard Rekomendasi")

    # A. LOAD CONFIG & TEMPLATES
    csv_path = os.path.join(CONFIG_DIR, "table_structure.csv")
    structure = []
    if os.path.exists(csv_path):
        with open(csv_path, "r", encoding="utf-8-sig", errors="replace") as f:
            # Auto-detect delimiter
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            try: dialect = sniffer.sniff(sample, delimiters=";")
            except: dialect = None
            reader = csv.DictReader(f, delimiter=dialect.delimiter if dialect else ";")
            structure = list(reader)

    # Get Metrics Columns (Same logic as dashboard endpoint)
    db_cols_info = con.execute("DESCRIBE master_data").fetchall()
    metadata_cols = {
        "valid_from", "valid_to", "commit_id", "source_file",
        "Provinsi", "Kabupaten/ Kota", "Kecamatan", 
        "Kode Wilayah Administrasi Desa", "Desa", "Status ID"
    }
    ordered_db_cols = [r[0] for r in db_cols_info if r[0] not in metadata_cols]

    # Load Narrative Templates
    item_names = [row.get("ITEM", "") for row in structure if row.get("ITEM")]
    templates = helpers_get_or_create_intervensi_kegiatan(item_names)

    # Call Helper
    calculated_rows = helpers_calculate_dashboard_stats(df_grid, structure, ordered_db_cols, templates)

    # B. SETUP COLUMNS & HEADERS
    # Set Widths (Approximating CSS pixels to Excel chars)
    # 40px ~= 5 chars, 300px ~= 40 chars, 1100px ~= 100 chars
    ws2.column_dimensions['A'].width = 5   # NO
    ws2.column_dimensions['B'].width = 15  # DIMENSI
    ws2.column_dimensions['C'].width = 15  # SUB
    ws2.column_dimensions['D'].width = 15  # INDIKATOR
    ws2.column_dimensions['E'].width = 45  # ITEM (Wide)
    for c in ['F','G','H','I','J','K']:
        ws2.column_dimensions[c].width = 8 # Scores
    ws2.column_dimensions['L'].width = 100 # INTERVENSI (Very Wide)

    # Pelaksana Columns
    for c in ['M','N','O','P','Q']:
        ws2.column_dimensions[c].width = 25

    # Write Headers (Rows 1 & 2)
    # Row 1: Merged Headers
    ws2.append(["NO", "DIMENSI", "SUB DIMENSI", "INDIKATOR", "ITEM", 
                "SKOR", "", "", "", "", "",         # Colspan 6
                "INTERVENSI",                       # Rowspan 2
                "PELAKSANA", "", "", "", ""])       # Colspan 5
    
    # Row 2: Sub Headers
    ws2.append(["", "", "", "", "", 
                "Rata-Rata", "1", "2", "3", "4", "5", 
                "", 
                "PUSAT", "PROVINSI", "KABUPATEN", "DESA", "LAINNYA"])

    # Merge Logic for Header
    ws2.merge_cells("A1:A2") # NO
    ws2.merge_cells("B1:B2") # DIM
    ws2.merge_cells("C1:C2") # SUB
    ws2.merge_cells("D1:D2") # IND
    ws2.merge_cells("E1:E2") # ITEM
    ws2.merge_cells("F1:K1") # SKOR Group
    ws2.merge_cells("L1:L2") # INTERVENSI
    ws2.merge_cells("M1:Q1") # PELAKSANA Group

    # Style Headers
    for r in [1, 2]:
        for cell in ws2[r]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = align_center

    # Write Data
    merge_tracker = {
        0: {"val": None, "start": 3},
        1: {"val": None, "start": 3},
        2: {"val": None, "start": 3},
        3: {"val": None, "start": 3}
    }
    start_row_idx = 3 # Excel starts at row 2 (row 1 is header)

    for i, row in enumerate(calculated_rows):
        current_excel_row = start_row_idx + i
        
        # Extract values from the dictionary returned by the helper
        row_values = [
            row.get("NO"), row.get("DIMENSI"), row.get("SUB DIMENSI"), row.get("INDIKATOR"), row.get("ITEM"),
            row.get("SKOR Rata-Rata"), row.get("SKOR 1"), row.get("SKOR 2"), row.get("SKOR 3"), row.get("SKOR 4"), row.get("SKOR 5"),
            row.get("INTERVENSI KEGIATAN"),
            row.get("PELAKSANA KEGIATAN PUSAT"), row.get("PELAKSANA KEGIATAN PROVINSI"), 
            row.get("PELAKSANA KEGIATAN KABUPATEN"), row.get("PELAKSANA KEGIATAN DESA"), 
            row.get("PELAKSANA KEGIATAN Lainnya")
        ]

        ws2.append(row_values)

        # Styling (Borders & Alignment)
        for col_idx, cell in enumerate(ws2[current_excel_row], 1):
            cell.border = thin_border
            cell.alignment = align_top # Top align is crucial for merged cells
            
            # Center align the Score columns (F to K -> 6 to 11)
            if 6 <= col_idx <= 11:
                cell.alignment = Alignment(horizontal='center', vertical='top')

        # Merging Logic
        # Iterate through mergeable columns (NO, DIMENSI, SUB, INDIKATOR)
        # Logic: If value changes OR parent value changed, trigger merge of previous block and start new.
        meta_vals = [row.get("NO"), row.get("DIMENSI"), row.get("SUB DIMENSI"), row.get("INDIKATOR")]
        for col_idx in [0, 1, 2, 3]: 
            val = meta_vals[col_idx]
            
            # Check if parent changed (e.g., Cant merge SUB DIMENSI if DIMENSI changed)
            parent_changed = any(meta_vals[p] != merge_tracker[p]["val"] for p in range(col_idx))
            
            if val != merge_tracker[col_idx]["val"] or parent_changed or (i == len(calculated_rows) - 1):
                # Merge the PREVIOUS block if it spanned more than 1 row
                prev_end = current_excel_row - 1 if not (i == len(calculated_rows) - 1) else current_excel_row
                prev_start = merge_tracker[col_idx]["start"]
                
                if prev_end > prev_start:
                    col_letter = get_column_letter(col_idx + 1)
                    ws2.merge_cells(f"{col_letter}{prev_start}:{col_letter}{prev_end}")
                
                # Reset tracker
                merge_tracker[col_idx]["val"] = val
                merge_tracker[col_idx]["start"] = current_excel_row

    # --- SHEET 3: DASHBOARD IKU ---
    ws3 = wb.create_sheet("Dashboard IKU")
    iku_csv_path = os.path.join(CONFIG_DIR, "table_structure_IKU.csv")
    iku_mapping_path = os.path.join(CONFIG_DIR, "iku_mapping.json")

    if os.path.exists(iku_csv_path) and os.path.exists(iku_mapping_path):
        with open(iku_mapping_path, "r", encoding="utf-8") as f:
            iku_mapping = json.load(f)

        with open(iku_csv_path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f, delimiter=';')
            rows = list(reader)

        if len(rows) >= 2:
            row1 = rows[0]
            row2 = rows[1]
            params_dict = params_dict or {}

            # Determine grouping hierarchy based on override or fallback to active filters
            group_col = params_dict.get("group_by", "")

            if not group_col:
                prov = params_dict.get("Provinsi", "")
                kab = params_dict.get("Kabupaten/ Kota", "")
                kec = params_dict.get("Kecamatan", "")

                # Since empty string now means "All selected/No filter", we check for empty
                if not kab or kab == "__NONE__":
                    group_col = "Provinsi"
                elif not kec or kec == "__NONE__":
                    group_col = "Kabupaten/ Kota"
                elif not params_dict.get("Desa"):
                    group_col = "Kecamatan"
                else:
                    group_col = "Desa"
            
            wilayah_header = group_col.upper()
            write_row1 = [wilayah_header if v == "WILAYAH" else v for v in row1]
            
            ws3.append(write_row1)
            ws3.append(row2)
            
            # Merge Row 1 headers horizontally
            current_val = write_row1[0]
            start_col = 1
            for col_idx in range(2, len(write_row1) + 1):
                v = write_row1[col_idx - 1]
                if v != "" and v != current_val:
                    if col_idx - 1 > start_col:
                        ws3.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col_idx-1)
                    current_val = v
                    start_col = col_idx
            if len(write_row1) > start_col:
                ws3.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=len(write_row1))
            
            # Merge Top-Left columns vertically
            for col_idx, v in enumerate(row1):
                if v in ["WILAYAH", "JLH DESA"]:
                    ws3.merge_cells(start_row=1, start_column=col_idx+1, end_row=2, end_column=col_idx+1)
            
            for r in [1, 2]:
                for cell in ws3[r]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = align_center

            valid_statuses = {"mandiri", "maju", "berkembang", "tertinggal", "sangat tertinggal"}
            col_idx_to_metric = {}
            parent_metrics = []
            parent_to_statuses = {}
            current_parent = ""

            for idx, p_val in enumerate(row1):
                if p_val.strip() != "":
                    current_parent = p_val.strip()
                    if current_parent not in parent_to_statuses:
                        parent_to_statuses[current_parent] = []
                    
                if idx >= 2:
                    sub_val = row2[idx].strip().lower()
                    col_idx_to_metric[idx] = {"parent": current_parent, "sub": sub_val}
                    if current_parent not in parent_metrics:
                        parent_metrics.append(current_parent)
                        
                    if sub_val in valid_statuses:
                        if sub_val not in parent_to_statuses[current_parent]:
                            parent_to_statuses[current_parent].append(sub_val)

            df_filtered = df_grid
            if df_filtered.height > 0 and group_col in df_filtered.columns:
                exprs = []
                for parent in parent_metrics:
                    children = iku_mapping.get(parent, [])
                    valid_children = [c.strip() for c in children if c.strip() in df_filtered.columns]
                    
                    if valid_children:
                        num_children = len(valid_children)
                        sum_expr = pl.sum_horizontal([pl.col(c).cast(pl.Float64, strict=False).fill_null(0) for c in valid_children])
                        exprs.append((sum_expr / num_children).alias(f"__iku_score_{parent}"))
                    else:
                        exprs.append(pl.lit(None).alias(f"__iku_score_{parent}"))
                
                if exprs:
                    df_filtered = df_filtered.with_columns(exprs)

                has_status = "Status ID" in df_filtered.columns
                agg_exprs = [pl.len().alias("JLH DESA")]
                for parent in parent_metrics:
                    agg_exprs.append(pl.col(f"__iku_score_{parent}").mean().alias(f"__iku_avg_{parent}"))
                    
                    if has_status:
                        for status in parent_to_statuses.get(parent, []):
                            safe_status = status.replace(" ", "_")
                            cond = (pl.col("Status ID") == status.upper()) & (pl.col(f"__iku_score_{parent}").fill_null(0) >= 4)
                            agg_exprs.append(cond.cast(pl.Int32).sum().alias(f"__iku_{safe_status}_{parent}"))
                    else:
                        for status in parent_to_statuses.get(parent, []):
                            safe_status = status.replace(" ", "_")
                            agg_exprs.append(pl.lit(0).alias(f"__iku_{safe_status}_{parent}"))

                df_grouped = df_filtered.group_by(group_col).agg(agg_exprs).sort(group_col)

                total_jlh_desa = df_grouped["JLH DESA"].sum()
                totals = {"JLH DESA": total_jlh_desa}
                for parent in parent_metrics:
                    totals[f"__iku_avg_{parent}"] = df_grouped[f"__iku_avg_{parent}"].sum()
                    for status in parent_to_statuses.get(parent, []):
                        safe_status = status.replace(" ", "_")
                        totals[f"__iku_{safe_status}_{parent}"] = df_grouped[f"__iku_{safe_status}_{parent}"].sum()

                total_row = ["TOTAL", total_jlh_desa]
                for idx in range(2, len(row2)):
                    metric_info = col_idx_to_metric.get(idx)
                    val_to_show = "-"
                    if metric_info:
                        parent = metric_info["parent"]
                        sub = metric_info["sub"]
                        t_total = sum(totals.get(f"__iku_{s.replace(' ', '_')}_{parent}", 0) for s in parent_to_statuses.get(parent, []))
                        
                        if sub == "rata-rata":
                            avg_val = totals.get(f"__iku_avg_{parent}")
                            if avg_val is not None: val_to_show = round(float(avg_val), 2)
                        elif sub in valid_statuses:
                            val_to_show = totals.get(f"__iku_{sub.replace(' ', '_')}_{parent}", 0)
                        elif sub == "total":
                            val_to_show = t_total
                        elif "capaian" in sub:
                            val_to_show = f"{(t_total / total_jlh_desa) * 100:.1f}%" if total_jlh_desa > 0 else "0.0%"
                    total_row.append(val_to_show)

                ws3.append(total_row)
                
                # Apply bold to TOTAL row and Number Format thousands
                for col_idx, cell in enumerate(ws3[3], 1):
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    if isinstance(cell.value, int):
                        cell.number_format = '#,##0'
                    elif isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'

                # Render Body Rows
                for row_data in df_grouped.to_dicts():
                    data_row = [row_data.get(group_col, "Unknown"), row_data.get("JLH DESA", 0)]
                    for idx in range(2, len(row2)):
                        metric_info = col_idx_to_metric.get(idx)
                        val_to_show = "-"
                        if metric_info:
                            parent = metric_info["parent"]
                            sub = metric_info["sub"]
                            t_total = sum(row_data.get(f"__iku_{s.replace(' ', '_')}_{parent}", 0) for s in parent_to_statuses.get(parent, []))
                            
                            if sub == "rata-rata":
                                avg_val = row_data.get(f"__iku_avg_{parent}")
                                if avg_val is not None: val_to_show = round(float(avg_val), 2)
                            elif sub in valid_statuses:
                                val_to_show = row_data.get(f"__iku_{sub.replace(' ', '_')}_{parent}", 0)
                            elif sub == "total":
                                val_to_show = t_total
                            elif "capaian" in sub:
                                jlh_desa = row_data.get("JLH DESA", 0)
                                val_to_show = f"{(t_total / jlh_desa) * 100:.1f}%" if jlh_desa > 0 else "0.0%"
                        data_row.append(val_to_show)
                    
                    ws3.append(data_row)
                    
                # Format Data Rows
                for r in range(4, ws3.max_row + 1):
                    for cell in ws3[r]:
                        cell.border = thin_border
                        if isinstance(cell.value, int):
                            cell.number_format = '#,##0'
                        elif isinstance(cell.value, float):
                            cell.number_format = '#,##0.00'

            ws3.column_dimensions['A'].width = 35
            ws3.column_dimensions['B'].width = 15

    return wb

def helpers_background_task_generate_pre_render_excel(year: str):
    """
    Background task that generates two complete master_data db to
    Excel (All data) files for the given year.
    
    Behavior:
    - Fetches ALL current active rows (valid_to IS NULL) ordered by ID_COL
    - Generates both versions via helpers_generate_excel_workbook:
        • Translated (scores → text recommendations)
        • Raw (original numeric scores)
    - Writes atomically using .tmp + os.replace to prevent Nginx from serving partial files
    - Skips silently if master_data table does not exist or contains zero rows
    - Logs progress with timestamps (flush=True for immediate Docker visibility)
    
    No return value.
    """
    start_time = time.time()
    # flush=True forces Docker to print immediately instead of buffering
    print(f"[{datetime.now()}]  Compiling background Master Excel for {year}...", flush=True)
    
    con, _ = helpers_get_db_connection(year)
    
    try:
        query = f'SELECT * EXCLUDE (valid_from, valid_to, commit_id, source_file) FROM master_data WHERE valid_to IS NULL ORDER BY "{ID_COL}" ASC'
        try:
            df_grid = con.execute(query).pl()
        except duckdb.CatalogException:
            print(f"[{datetime.now()}]  Table not built yet for {year}. Skipping.", flush=True)
            return # Skip if table not built yet
        
        if df_grid.height == 0: 
            print(f"[{datetime.now()}]  No data found for {year}. Skipping.", flush=True)
            return

        print(f"[{datetime.now()}]  Data fetched ({df_grid.height} rows). Building Excel... (This may take several minutes)", flush=True)

        # Call the new clean helper
        wb_trans = helpers_generate_excel_workbook(con, df_grid, do_translate=True, params_dict={})
        file_path_trans = os.path.join(EXPORT_FOLDER, f"Export_Nasional_{year}_text.xlsx")
        temp_trans = f"{file_path_trans}.tmp"
        
        print(f"[{datetime.now()}] Saving Translated file to disk...", flush=True)

        # Save to a temporary file first to prevent Nginx from serving an incomplete file
        wb_trans.save(temp_trans)
        if os.path.exists(temp_trans) and os.path.getsize(temp_trans) > 0:
            os.replace(temp_trans, file_path_trans) # Atomic swap
        # Explicitly free translated workbook before building raw
        del wb_trans
        gc.collect()

        # Compile Raw Version
        wb_raw = helpers_generate_excel_workbook(con, df_grid, do_translate=False, params_dict={})
        file_path_raw = os.path.join(EXPORT_FOLDER, f"Export_Nasional_{year}_skor.xlsx")
        temp_raw = f"{file_path_raw}.tmp"
        
        # Explicit validation checks
        print(f"[{datetime.now()}] Saving Raw file to disk...", flush=True)
        wb_raw.save(temp_raw)
        if os.path.exists(temp_raw) and os.path.getsize(temp_raw) > 0:
            os.replace(temp_raw, file_path_raw) # Atomic swap

        # Free everything after done
        del wb_raw
        del df_grid
        gc.collect()

        elapsed = round(time.time() - start_time, 2)
        print(f"[{datetime.now()}] ✅ Successfully compiled BOTH Translated & Raw files in {elapsed}s", flush=True)

    except Exception as e:
        print(f"[{datetime.now()}] ❌ Failed to generate master excel: {e}", flush=True)
        traceback.print_exc() # Prints the exact line that crashed to Docker logs
    finally:
        con.close()
        # Cleanup any stuck temp files
        for t in [f"{os.path.join(EXPORT_FOLDER, f'Export_Nasional_{year}_text.xlsx')}.tmp", 
                  f"{os.path.join(EXPORT_FOLDER, f'Export_Nasional_{year}_skor.xlsx')}.tmp"]:
            if os.path.exists(t):
                try: os.remove(t)
                except: pass

def helpers_get_or_create_intervensi_kegiatan(items: list[str]):
    """
    Loads or creates intervensi_kegiatan_mapping.json.
    If file doesn't exist, generates default templates based on rekomendasi.json (when available).
    """
    # Try Load
    if os.path.exists(INTERVENTION_FILE):
        try:
            with open(INTERVENTION_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except: pass # Fallback if corrupt

    # Create Defaults (using rekomendasi.json logic if available, else generic)
    defaults = {}

    # Try to load existing rekomendasis to pre-seed
    recs = {}
    rec_path = os.path.join(CONFIG_DIR, "rekomendasi.json")
    if os.path.exists(rec_path):
        try:
            with open(rec_path, "r", encoding="utf-8") as f: recs = json.load(f)
        except: pass

    for item in items:
        # If we have specific logic in rekomendasi.json, map it. 
        # Otherwise create a generic placeholder.
        if item in recs:
            defaults[item] = recs[item]
        else:
            # Generic Fallback Template that is WRONG. 
            # YOU NEED TO EDIT THE TEMPLATE
            # this below is when rekomendasi.json isn't filled
            defaults[item] = {
                "1": f"Perlu peningkatan {item} (Sangat Kurang)",
                "2": f"Perlu peningkatan {item} (Kurang)",
                "3": f"Perlu peningkatan {item} (Cukup)",
                "4": f"Perlu peningkatan {item} (Baik)",
                "5": None
            }

    # Save to Disk
    try:
        with open(INTERVENTION_FILE, "w", encoding="utf-8") as f:
            json.dump(defaults, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ Could not save intervensi_kegiatan: {e}")

    return defaults

# reusable dashboard calculate
def helpers_calculate_dashboard_stats(df_data, structure, ordered_db_cols, templates):
    """
    Calculates dashboard statistics (average score, count per score level, and intervention narrative)
    for each row in table_structure.csv using the provided DataFrame.
    Shared logic used by both UI rendering and Excel export.
    """
    calculated_rows = []

    for idx, row in enumerate(structure):
        # Default Empty Stats
        stats = {
            "SKOR Rata-Rata": "",
            "SKOR 1": "", "SKOR 2": "", "SKOR 3": "", "SKOR 4": "", "SKOR 5": "",
            "INTERVENSI KEGIATAN": ""
        }

        # Calculation Logic (Match CSV Row Index -> DB Column Index)
        if idx < len(ordered_db_cols):
            target_col = ordered_db_cols[idx]

            if target_col in df_data.columns:
                try:
                    # Select specific column as Series
                    s = df_data.select(pl.col(target_col).cast(pl.Int64, strict=False)).to_series()
                    
                    avg = s.mean()
                    c1 = (s == 1).sum()
                    c2 = (s == 2).sum()
                    c3 = (s == 3).sum()
                    c4 = (s == 4).sum()
                    c5 = (s == 5).sum()

                    if avg is not None: 
                        stats["SKOR Rata-Rata"] = round(avg, 2)
                    if c1: stats["SKOR 1"] = c1
                    if c2: stats["SKOR 2"] = c2
                    if c3: stats["SKOR 3"] = c3
                    if c4: stats["SKOR 4"] = c4
                    if c5: stats["SKOR 5"] = c5

                    # Narrative Generation
                    narrative_parts = []
                    t_map = templates.get(row.get("ITEM", ""), {})
                    cnt_map = {1: c1, 2: c2, 3: c3, 4: c4, 5: c5}
                    
                    for score in [1, 2, 3, 4, 5]:
                        if cnt_map[score] > 0 and t_map.get(str(score)):
                            # Format number with commas for text
                            narrative_parts.append(f"{cnt_map[score]:,} {t_map[str(score)]}")
                    
                    if narrative_parts:
                        stats["INTERVENSI KEGIATAN"] = "\n".join(narrative_parts)

                except Exception as e:
                    print(f"Calc error on column {target_col}: {e}")

        # Merge Stats into the Original Row
        # This ensures we have the metadata (NO, DIMENSI...) + The Calculated Stats + The Pelaksana info
        row.update(stats)
        calculated_rows.append(row)

    return calculated_rows

# rendering dashboard
# the logic for rendering moved here, reason so no one can inspect the logic on <script>
def helpers_render_dashboard_html(calculated_rows: list[dict]) -> str:
    """
    Server-side HTML renderer for the dashboard table.
    - Builds full <thead> + <tbody> with proper rowspans for merged columns (NO, DIMENSI, SUB DIMENSI, INDIKATOR)
    - Escapes content to prevent XSS
    - Adds data-col-idx and resizer elements for frontend features (freeze, resize)
    """
    if not calculated_rows:
        return "<tbody><tr><td colspan='17' class='text-center p-4'>No Data</td></tr></tbody>"

    # Define Column Indices mapping to ensure alignment between Header and Body
    # 0:NO, 1:DIM, 2:SUB, 3:IND, 4:ITEM, 5-10:SKOR, 11:INTERVENSI, 12-16:PELAKSANA
    
    # --- BUILD HEADER (Static Structure based on your CSV layout) ---
    #add 'data-col-idx' and the 'resizer' div.
    html = """
    <thead>
        <tr>
            <th class="col-no relative" rowspan="2" data-col-idx="0"><span>NO</span><div class="resizer"></div></th>
            <th class="col-dimensi relative" rowspan="2" data-col-idx="1"><span>DIMENSI</span><div class="resizer"></div></th>
            <th class="col-sub_dimensi relative" rowspan="2" data-col-idx="2"><span>SUB DIMENSI</span><div class="resizer"></div></th>
            <th class="col-indikator relative" rowspan="2" data-col-idx="3"><span>INDIKATOR</span><div class="resizer"></div></th>
            <th class="col-item relative" rowspan="2" data-col-idx="4"><span>ITEM</span><div class="resizer"></div></th>
            <th colspan="6">SKOR</th>
            <th class="col-intervensi relative" rowspan="2" data-col-idx="11">INTERVENSI<div class="resizer"></div></th>
            <th colspan="5">PELAKSANA</th>
        </tr>
        <tr>
            <th class="col-score relative" data-col-idx="5">Rata-Rata<div class="resizer"></div></th>
            <th class="col-score relative" data-col-idx="6">1<div class="resizer"></div></th>
            <th class="col-score relative" data-col-idx="7">2<div class="resizer"></div></th>
            <th class="col-score relative" data-col-idx="8">3<div class="resizer"></div></th>
            <th class="col-score relative" data-col-idx="9">4<div class="resizer"></div></th>
            <th class="col-score relative" data-col-idx="10">5<div class="resizer"></div></th>
            <th class="col-pusat relative" data-col-idx="12">PUSAT<div class="resizer"></div></th>
            <th class="col-prov relative" data-col-idx="13">PROV<div class="resizer"></div></th>
            <th class="col-kab relative" data-col-idx="14">KAB<div class="resizer"></div></th>
            <th class="col-desa relative" data-col-idx="15">DESA<div class="resizer"></div></th>
            <th class="col-lain relative" data-col-idx="16">LAIN<div class="resizer"></div></th>
        </tr>
    </thead>
    <tbody>
    """

    # --- CALCULATE ROW SPANS (Ported from JS) ---
    # pre-calculate spans to know when to output <td rowspan="X"> or skip
    n_rows = len(calculated_rows)
    row_spans = [{"NO": 0, "DIMENSI": 0, "SUB DIMENSI": 0, "INDIKATOR": 0} for _ in range(n_rows)]
    merge_keys = ["NO", "DIMENSI", "SUB DIMENSI", "INDIKATOR"]

    # Map keys to their CSS classes
    col_css_map = {
        "NO": "col-no",
        "DIMENSI": "col-dimensi",
        "SUB DIMENSI": "col-sub_dimensi",
        "INDIKATOR": "col-indikator"
    }

    for k in merge_keys:
        i = 0
        while i < n_rows:
            # Start of a potential span
            span = 1
            j = i + 1
            while j < n_rows:
                # Check if current row matches start row
                if calculated_rows[i].get(k) == calculated_rows[j].get(k):
                    # (e.g., Can't merge "SUB DIMENSI A" if "DIMENSI" changed)
                    parents_match = True
                    idx_k = merge_keys.index(k)
                    for p_idx in range(idx_k): # Check all keys before current K
                        p_key = merge_keys[p_idx]
                        if calculated_rows[i].get(p_key) != calculated_rows[j].get(p_key):
                            parents_match = False
                            break
                    
                    if parents_match:
                        span += 1
                        row_spans[j][k] = -1 # Mark as "skip"
                        j += 1
                    else:
                        break
                else:
                    break
            
            row_spans[i][k] = span
            i = j # Jump forward

    # --- BUILD BODY HTML ---
    for i, row in enumerate(calculated_rows):
        html += "<tr>"
        
        # Helper
        def make_td(content, idx, css_class="", rowspan=1, is_merged=False):
            # Combine passed class with merged-cell class
            cls = f"{css_class} {'merged-cell' if is_merged else ''}".strip()
            
            # Default padding is 6px, but 'col-no' has !important padding:0 in CSS
            # add vertical-align: top to everything.
            style = "vertical-align: top;"
            
            # Only add whitespace wrap if it's NOT the NO column (which is centered/tight)
            if "col-no" not in css_class:
                style += " white-space: pre-wrap; padding: 6px;"
            
            # File: middleware.py (Function: helpers_render_dashboard_html)
            # Vulnerability: You use f-strings to build HTML: f'>{content or ""}</td>'.
            # The Risk: If an attacker uploads an Excel file where the "Desa"
            # column contains <script>alert('Hacked')</script>,
            # that script will execute in the browser of every admin who views the dashboard.
            safe_content = h.escape(str(content)) if content else ""
            return f'<td class="{cls}" data-col-idx="{idx}" rowspan="{rowspan}" style="{style}">{safe_content}</td>'

        # 0-3: Merged Columns (NO, DIM, SUB, IND)
        for idx, key in enumerate(merge_keys): 
            span = row_spans[i][key]
            if span != -1:
                # FIX: Pass the specific CSS class (e.g., 'col-no')
                html += make_td(row.get(key), idx, css_class=col_css_map.get(key, ""), rowspan=span, is_merged=(span > 1))

        # 4: ITEM
        html += make_td(row.get("ITEM"), 4, css_class="col-item")

        # 5-10: SCORES
        score_keys = ["SKOR Rata-Rata", "SKOR 1", "SKOR 2", "SKOR 3", "SKOR 4", "SKOR 5"]
        for idx, key in enumerate(score_keys, 5):
            val = row.get(key)
            if key == "SKOR Rata-Rata":
                display = str(val) if val is not None else ""
            html += f'<td class="col-score" data-col-idx="{idx}" style="text-align:center;">{display}</td>'

        # 11: INTERVENSI
        html += make_td(row.get("INTERVENSI KEGIATAN"), 11, css_class="col-intervensi")

        # 12-16: PELAKSANA
        pelaksana_keys = ["PUSAT", "PROVINSI", "KABUPATEN", "DESA", "Lainnya"]
        pelaksana_css = ["col-pusat", "col-prov", "col-kab", "col-desa", "col-lain"]
        for j, suffix in enumerate(pelaksana_keys):
            idx = 12 + j
            key = f"PELAKSANA KEGIATAN {suffix}"
            html += make_td(row.get(key), idx, css_class=pelaksana_css[j])

        html += "</tr>"

    html += "</tbody>"
    return html

# IKU Dashboard rendering
def helpers_render_iku_dashboard(df_filtered: pl.DataFrame, params_dict: dict) -> str:
    """
    Renders HTML table for IKU Dashboard from filtered DataFrame.
    
    Logic:
    - Loads table_structure_IKU.csv (semicolon-delimited) and iku_mapping.json
    - Determines grouping level from filters (Provinsi → Kabupaten → Kecamatan → Desa)
    - Maps CSV headers: parents (merged colspans) + subs (statuses, rata-rata, total, capaian)
    - Computes per-parent IKU scores: averages over mapped children columns
    - Aggregates by group: JLH DESA, averages, status counts (≥4 threshold)
    - Adds TOTAL row with sums/averages
    - Applies heatmaps: green intensity for data, red-green hue for capaian %
    - Escapes wilayah names
    - Handles missing columns/files with error rows
    
    Returns:
        str: full <thead> + <tbody> HTML (with data-col-idx, resizers, classes for styling)
    """
    iku_csv_path = os.path.join(CONFIG_DIR, "table_structure_IKU.csv")
    iku_mapping_path = os.path.join(CONFIG_DIR, "iku_mapping.json")

    if not os.path.exists(iku_csv_path):
        return "<tbody><tr><td class='p-3 text-center'>Error: table_structure_IKU.csv missing</td></tr></tbody>"

    # Load mapping
    iku_mapping = {}
    if os.path.exists(iku_mapping_path):
        with open(iku_mapping_path, "r", encoding="utf-8") as f:
            iku_mapping = json.load(f)

    with open(iku_csv_path, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f, delimiter=';')
        rows = list(reader)

    if len(rows) < 2:
        return "<tbody><tr><td class='p-4 text-center'>Error: Invalid table format</td></tr></tbody>"

    row1 = rows[0]
    row2 = rows[1]

    # Determine grouping hierarchy based on override or fallback to active filters
    group_col = params_dict.get("group_by", "")
    
    if not group_col:
        prov = params_dict.get("Provinsi", "")
        kab = params_dict.get("Kabupaten/ Kota", "")
        kec = params_dict.get("Kecamatan", "")

        # Since empty string now means "All selected/No filter", we check for empty
        if not kab or kab == "__NONE__":
            group_col = "Provinsi"
        elif not kec or kec == "__NONE__":
            group_col = "Kabupaten/ Kota"
        elif not params_dict.get("Desa"):
            group_col = "Kecamatan"
        else:
            group_col = "Desa"

    # Make the table header explicitly show the level you are viewing
    wilayah_header = group_col.upper()

    # Define valid statuses that can exist in the DB
    valid_statuses = {"mandiri", "maju", "berkembang", "tertinggal", "sangat tertinggal"}

    # --- MAP CSV HEADERS ---
    col_idx_to_metric = {}
    parent_metrics = []
    parent_to_statuses = {}
    
    # Track the current parent to handle merged cells or repeated headers seamlessly
    current_parent = ""
    for idx, p_val in enumerate(row1):
        if p_val.strip() != "":
            current_parent = p_val.strip()
            if current_parent not in parent_to_statuses:
                parent_to_statuses[current_parent] = []
            
        if idx >= 2: # Skip WILAYAH and JLH DESA
            sub_val = row2[idx].strip().lower()
            col_idx_to_metric[idx] = {
                "parent": current_parent,
                "sub": sub_val
            }
            if current_parent not in parent_metrics:
                parent_metrics.append(current_parent)
                
            # Track which specific statuses this parent uses for dynamic Totals
            if sub_val in valid_statuses:
                if sub_val not in parent_to_statuses[current_parent]:
                    parent_to_statuses[current_parent].append(sub_val)

    # --- Build HTML Header ---
    html = "<thead><tr>"
    
    colspans = []
    if row1:
        current_val = row1[0]
        count = 1
        for val in row1[1:]:
            if val == current_val and val != "":
                count += 1
            else:
                colspans.append((current_val, count))
                current_val = val
                count = 1
        colspans.append((current_val, count))

    # Top level headers
    col_idx_tracker = 0
    for val, span in colspans:
        if val == "WILAYAH":
            html += (
                f'<th rowspan="2" class="relative bg-gray-200 dark:bg-slate-700 border dark:border-slate-600 '
                f'p-2 iku-header-wilayah" data-col-idx="{col_idx_tracker}">'
                f'<span>{wilayah_header}</span><div class="resizer"></div></th>'
            )
            col_idx_tracker += 1
        elif val == "JLH DESA":
            html += (
                f'<th rowspan="2" class="relative bg-gray-200 dark:bg-slate-700 border dark:border-slate-600 '
                f'p-2 iku-header-jlh" data-col-idx="{col_idx_tracker}">'
                f'<span>{val}</span><div class="resizer"></div></th>'
            )
            col_idx_tracker += 1
        else:
            html += (
                f'<th colspan="{span}" class="text-center relative bg-gray-200 dark:bg-slate-700 border '
                f'dark:border-slate-600 p-3 iku-header-group" style="width: 300px;min-width: 300px;">'
                f'<span>{val}</span><div class="resizer"></div></th>'
            )

    html += "</tr><tr>"

    # Sub level headers (Starting from data-col-idx=2)
    for idx, val in enumerate(row2):
        if idx < len(row1) and row1[idx] not in ["WILAYAH", "JLH DESA"]:
            html += (
                f'<th class="relative text-center bg-gray-100 dark:bg-slate-800 border dark:border-slate-600 '
                f'p-3 text-xs iku-header-sub" data-col-idx="{idx}">'
                f'<span>{val}</span><div class="resizer"></div></th>'
            )

    html += "</tr></thead><tbody>"

    # --- Build HTML Body ---
    if df_filtered.height > 0:
        if group_col in df_filtered.columns:
            # Calculate Desa-Level Averages for each Parent Metric
            exprs = []
            for parent in parent_metrics:
                children = iku_mapping.get(parent, [])
                valid_children = [c.strip() for c in children if c.strip() in df_filtered.columns]
                
                if valid_children:
                    num_children = len(valid_children)
                    # Sum horizontally, treating missing/null as 0
                    sum_expr = pl.sum_horizontal([pl.col(c).cast(pl.Float64, strict=False).fill_null(0) for c in valid_children])
                    avg_expr = sum_expr / num_children
                    exprs.append(avg_expr.alias(f"__iku_score_{parent}"))
                else:
                    # If columns don't exist in DB, assign None
                    exprs.append(pl.lit(None).alias(f"__iku_score_{parent}"))
            
            # Apply to dataframe
            if exprs:
                df_filtered = df_filtered.with_columns(exprs)

            # STEP 2: Group by Region (Wilayah) and aggregate
            has_status = "Status ID" in df_filtered.columns
            agg_exprs = [pl.len().alias("JLH DESA")]
            for parent in parent_metrics:
                agg_exprs.append(pl.col(f"__iku_score_{parent}").mean().alias(f"__iku_avg_{parent}"))
                
                if has_status:
                    # Dynamically count ONLY the statuses required by this parent's CSV headers
                    for status in parent_to_statuses.get(parent, []):
                        safe_status = status.replace(" ", "_")
                        cond = (pl.col("Status ID") == status.upper()) & (pl.col(f"__iku_score_{parent}").fill_null(0) >= 4)
                        agg_exprs.append(cond.cast(pl.Int32).sum().alias(f"__iku_{safe_status}_{parent}"))
                else:
                    for status in parent_to_statuses.get(parent, []):
                        safe_status = status.replace(" ", "_")
                        agg_exprs.append(pl.lit(0).alias(f"__iku_{safe_status}_{parent}"))

            # Group by hierarchy
            df_grouped = df_filtered.group_by(group_col).agg(agg_exprs).sort(group_col)

            # Calculate TOTAL Row & Column Maximums for Heatmap
            total_jlh_desa = df_grouped["JLH DESA"].sum()
            totals = {"JLH DESA": total_jlh_desa}
            col_maxes = {}

            for parent in parent_metrics:
                totals[f"__iku_avg_{parent}"] = df_grouped[f"__iku_avg_{parent}"].sum()
                col_maxes[f"__iku_avg_{parent}"] = df_grouped[f"__iku_avg_{parent}"].max()

                parent_total_series = pl.Series(name="tmp", values=[0] * df_grouped.height)

                for status in parent_to_statuses.get(parent, []):
                    safe_status = status.replace(" ", "_")
                    totals[f"__iku_{safe_status}_{parent}"] = df_grouped[f"__iku_{safe_status}_{parent}"].sum()
                    col_maxes[f"__iku_{safe_status}_{parent}"] = df_grouped[f"__iku_{safe_status}_{parent}"].max()
                    
                    # Accumulate for the parent's TOTAL heatmap
                    parent_total_series = parent_total_series + df_grouped[f"__iku_{safe_status}_{parent}"]
                    
                col_maxes[f"__iku_total_{parent}"] = parent_total_series.max()

            # Inject TOTAL Row HTML (Forced to precisely match header classes, no tailwind text coloring)
            html += "<tr class='font-extrabold'>"
            html += (
                f'<td class="p-3 border dark:border-slate-600 bg-gray-200 dark:bg-slate-700 text-slate-800'
                f'dark:text-slate-100 iku-cell-wilayah" data-col-idx="0">TOTAL</td>'
            )
            html += (
                f'<td class="p-3 border dark:border-slate-600 bg-gray-200 dark:bg-slate-700 text-center'
                f'font-mono iku-cell-jlh" data-col-idx="1">{total_jlh_desa:,}</td>'
            )

            for idx in range(2, len(row2)):
                metric_info = col_idx_to_metric.get(idx)
                val_to_show = "-"
                
                if metric_info:
                    parent = metric_info["parent"]
                    sub = metric_info["sub"]
                    
                    # Dynamically sum all valid statuses mapped to THIS parent
                    t_total = sum(totals.get(f"__iku_{s.replace(' ', '_')}_{parent}", 0) for s in parent_to_statuses.get(parent, []))
                    
                    if sub == "rata-rata":
                        avg_val = totals.get(f"__iku_avg_{parent}")
                        if avg_val is not None:
                            val_to_show = f"{avg_val:.2f}"
                    elif sub in valid_statuses:
                        safe_sub = sub.replace(" ", "_")
                        val = totals.get(f"__iku_{safe_sub}_{parent}", 0)
                        val_to_show = f"{val:,}"
                    elif sub == "total":
                        val_to_show = f"{t_total:,}"
                    elif "capaian" in sub:
                        if total_jlh_desa > 0:
                            val_to_show = f"{(t_total / total_jlh_desa) * 100:.1f}%"
                        else:
                            val_to_show = "0.0%"

                html += (
                    f'<td class="p-3 border dark:border-slate-600 bg-gray-200 dark:bg-slate-700 text-center '
                    f'iku-cell-data text-slate-800 dark:text-slate-100" data-col-idx="{idx}">{val_to_show}</td>'
                )
            html += "</tr>"
            
            # Inject calculated values into HTML body (With CSS Heatmap logic applied)
            for row_data in df_grouped.to_dicts():
                html += "<tr class='hover:bg-gray-50 dark:hover:bg-slate-800/50 transition-colors'>"
                wilayah = row_data.get(group_col, "Unknown")
                jlh_desa = row_data.get("JLH DESA", 0)
                
                # Wilayah & Jumlah Desa Columns
                html += (
                    f'<td class="p-3 border dark:border-slate-700 font-bold text-slate-700 dark:text-slate-200 '
                    f'iku-cell-wilayah" data-col-idx="0">{h.escape(str(wilayah))}</td>'
                )
                html += (
                    f'<td class="p-3 border dark:border-slate-700 text-center font-mono iku-cell-jlh" '
                    f'data-col-idx="1">{jlh_desa:,}</td>'
                )

                # Map corresponding calculated Data
                for idx in range(2, len(row2)):
                    metric_info = col_idx_to_metric.get(idx)
                    val_to_show = "-"
                    inline_style = "" # Standard transparent default
                    
                    if metric_info:
                        parent = metric_info["parent"]
                        sub = metric_info["sub"]
                        
                        # Dynamically sum all valid statuses mapped to THIS parent
                        t_total = sum(row_data.get(f"__iku_{s.replace(' ', '_')}_{parent}", 0) for s in parent_to_statuses.get(parent, []))
                        
                        raw_val = 0
                        max_val = 1
 
                        if sub == "rata-rata":
                            avg_val = row_data.get(f"__iku_avg_{parent}")
                            if avg_val is not None:
                                val_to_show = f"{avg_val:.2f}"
                                raw_val = avg_val
                                max_val = col_maxes.get(f"__iku_avg_{parent}", 1)
                        elif sub in valid_statuses:
                            safe_sub = sub.replace(" ", "_")
                            val = row_data.get(f"__iku_{safe_sub}_{parent}", 0)
                            val_to_show = f"{val:,}"
                            raw_val = val
                            max_val = col_maxes.get(f"__iku_{safe_sub}_{parent}", 1)
                        elif sub == "total":
                            val_to_show = f"{t_total:,}"
                            raw_val = t_total
                            max_val = col_maxes.get(f"__iku_total_{parent}", 1)
                        
                        # Apply Background Heatmap Calculation
                        if "capaian" in sub:
                            if jlh_desa > 0:
                                ratio = t_total / jlh_desa
                                val_to_show = f"{ratio * 100:.1f}%"
                                # Capaian heatmap: Hue ranges from 0 (Red) to 120 (Green)
                                hue = int(ratio * 120)
                                inline_style = f"background-color: hsla({hue}, 70%, 50%, 0.4);"
                            else:
                                val_to_show = "0.0%"
                                inline_style = "background-color: hsla(0, 70%, 50%, 0.4);"
                        else:
                            # Standard Data heatmap: Alpha transparency scales green intensity
                            if max_val and max_val > 0:
                                intensity = raw_val / max_val
                                # Cap opacity at 0.55 so text stays readable
                                inline_style = f"background-color: rgba(34, 197, 94, {min(1.0, max(0.0, intensity)) * 0.55});"

                    html += (
                        f'<td class="p-3 border dark:border-slate-700 text-center font-semibold text-slate-800 dark:text-slate-100 '
                        f'iku-cell-data" style="padding-top: 2px; padding-bottom: 2px; {inline_style}" data-col-idx="{idx}">{val_to_show}</td>'
                    )
                html += "</tr>"
        else:
            html += (
                f"<tr><td colspan='{len(row2)}' class='text-center p-4 text-red-500'>"
                f"Kolom '{group_col}' tidak ditemukan di database</td></tr>"
            )
    else:
        html += f"<tr><td colspan='{len(row2)}' class='text-center p-4'>Tidak ada data untuk filter ini</td></tr>"

    html += "</tbody>"
    return html
